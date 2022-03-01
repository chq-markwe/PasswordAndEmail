import tkinter as tk
from tkinter.filedialog import askopenfilename
from tkinter.filedialog import asksaveasfilename
import pandas as pd
import win32com.client
import random
import string
import openpyxl

# --- classes ---

class MyWindow:
    def __init__(self, parent):

        self.parent = parent
        self.parent.title('Password Generator and Email Sender')
        self.parent.iconbitmap('flag.ico')

        self.filename = None
        self.df = None

        self.text = tk.Text(self.parent)
        self.text.pack()
        self.text.insert('insert', '\n'+ '  Instructions for how to use this Utility' + '\n' + '\n' \
        + '  *  Click on the Generate Passwords button'+ '\n' + '\n'\
        + '  *  Select the file in the windows explorer' + '\n' + '\n'\
        + '  *  Click Open' + '\n'+ '\n'\
        + '  *  Select the location and name of the Save file' + '\n' + '\n'\
        + '      * You should see "Passwords Generated" display here' + '\n'+ '\n'\
        + '  When passwords generated have been updated in ECCO run this utility again' + '\n'+ '\n'\
        + '  *  Select the Send Emails button'+ '\n'+ '\n'\
        + '  *  Select the file saved in the previous step and click Open' + '\n'+ '\n'\
        + '      * You should see "Emails Sent" display here' + '\n' + '\n'+ '\n'\
        + '  ***Note: You may have to scroll past the displayed data in order to see'+ '\n' \
        + '  the confirmation messages'+ '\n' + '\n'+ '\n')
        # Generate Passwords button
        self.script_button = tk.Button(self.parent, text='Generate Passwords', command=self.generate_passwords)
        self.script_button.pack()

        # Generate Emails button
        self.script_button = tk.Button(self.parent, text='Send Emails', command=self.send_emails)
        self.script_button.pack()

        # Exit button
        self.script_button = tk.Button(self.parent, text="Exit", command=self.parent.destroy)
        self.script_button.pack(pady=20)

    def generate_passwords(self):
        def generate_password():

            length = 9
            symbol_list = r"""#%&!@?*$"""
            lower = string.ascii_lowercase
            upper = string.ascii_uppercase
            num = string.digits
            symbols = random.sample(symbol_list, 1)
            first = lower + upper + num

            temp = random.sample(first, length)

            # create the password
            password = "".join(temp + symbols)

            # return the password
            return (password)

        # Ask for input file
        name = askopenfilename(filetypes=[('Excel', ('*.xls', '*.xslm', '*.xlsx')), ('CSV', '*.csv',)])

        if name:
            if name.endswith('.csv'):
                self.df = pd.read_csv(name)
            else:
                self.df = pd.read_excel(name)

            self.filename = name

            # display directly
            self.text.insert('end', str(self.df.head()) + '\n')

        wb = openpyxl.load_workbook(name)
        wb.sheetnames  # get names of all spreadsheet in the file
        ws = wb["Sheet1"]  # get the first spreadsheet by name
        ws.max_row  # get the number of rows in the sheet

        for i in range(2, ws.max_row + 1):
            pw = generate_password()
            ws[f"F{i}"].value = pw

        # Ask for saved file name
        fname = asksaveasfilename(filetypes=(("Excel files", "*.xlsx"),
                                        ("All files", "*.*")))
        # note: this will fail unless user ends the fname with ".xlsx"
        # self.df.to_excel(fname)
        wb.save(fname)

        self.filename = fname
        self.df = pd.read_excel(fname)

        self.text.insert('end', self.filename + '\n')
        self.text.insert('end', str(self.df.head()) + '\n')

        self.text.insert('insert', '\n'+ 'Passwords Generated'+ '\n'+ '\n')


    def send_emails(self):

        def send_email():
            # Generate the email
            outlook = win32com.client.Dispatch('outlook.application')

            # choose sender account
            send_account = None
            for account in outlook.Session.Accounts:
                if account.DisplayName == 'eccosupport@expeditors.com':
                    send_account = account
                    break

            # reading the spreadsheet
            email_list = askopenfilename(filetypes=[('Excel', ('*.xls', '*.xslm', '*.xlsx')), ('CSV', '*.csv',)])

            if email_list:
                if email_list.endswith('.csv'):
                    self.df = pd.read_csv(email_list)
                else:
                    self.df = pd.read_excel(email_list)

                self.filename = email_list

            # getting the info from the spreadsheet
            first_names = self.df['FIRST NAME'].tolist()
            usernames = self.df['USERNAME'].tolist()
            emails = self.df['EMAIL'].tolist()
            passwords = self.df['PASSWORD'].tolist()

            # iterate through the records
            for i in range(len(emails)):
                mail = outlook.CreateItem(0)

                # Code to assign the sending account
                mail._oleobj_.Invoke(*(64209, 0, 8, 0, send_account))

                # for every record get the name and the email addresses
                first_name = first_names[i]
                username = usernames[i]
                email = emails[i]
                password = passwords[i]

                # the message to be emailed
                mail.To = email
                mail.Subject = f'User Account Created for {username}'
                mail.HTMLBody = (f"""Hello {first_name},<br><br>
                     Your account has been created.<br><br>
                     Please see below for your username and temporary password:<br><br>
                     Username: {username}<br>
                     Password: {password}<br><br>
                     Please note: you will have to change your password upon logging in to ECCO<br><br><br>
                     Best Regards, <br><br>
                     <strong>Critical Logistics Management</strong>
                     <strong><p style='color:red;font-size:.7rem'>Expeditors Carrier Capacity Optimization</p><strong><br><br>
                     <strong>Email</strong> eccosupport@expeditors.com <br><br>
                     <a href='https://www.expeditors.com'>
                     <img src ='https://info.expeditors.com/hubfs/without%20tag%20line.gif', alt = 'Expeditors Logo no tag', width = '150px'> 
                     </a>
                     """)
                mail.BCC = 'eccosupport@expeditors.com'

                mail.Send()

        send_email()

        self.text.insert('insert', '\n' + 'Emails Sent')

# --- main ---

if __name__ == '__main__':
    root = tk.Tk()
    top = MyWindow(root)
    root.mainloop()